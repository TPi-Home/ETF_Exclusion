#My intentions are to search through reddit user nobjos's excel sheet for individual stocks,
#thus ruling out market indexes, mutual funds and ETFs. I am going to compartmentalize this
#with different sheets for each exclusionary string. 
#remove ishares/QQQ/ETF

#REQUEST FOR EDITORS: IF YOU SEE ANY INDEX TYPE OF ENTRY, PLEASE DELETE
from openpyxl import load_workbook 
from openpyxl import Workbook
from openpyxl.utils import range_boundaries
from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
from copy import copy

#INSERT GET USER INPUT HERE

wb= load_workbook(filename="Trades_Sheet.xlsx", data_only=True)
ws = wb["All_Trades"]
wsnew = wb.create_sheet("New")
#This will be a method to call on so I can delete the entire row and now just the contents of one
for row in ws.rows:
    for cell in row:
        if "ETF" in str(cell.value):
            print("Deleted:", cell.value)
            r= cell.row
            wsnew.delete_rows(row[0].row,1)
        elif "QQQ" in str(cell.value):
            print("Deleted:", cell.value)
            r= cell.row
            wsnew.delete_rows(row[0].row,1)
        elif "iShares" in str(cell.value):
            print("Deleted:", cell.value)
            r= cell.row
            print(r)
            wsnew.delete_rows(row[0].row,1)
        else:
            wsnew[cell.coordinate]=cell.value

#function to iterate and fully delete rows
#for row in ws.rows:
#    for cell in row:
#        if cell.value is None:
#            print(cell)
#            r= cell.row
#            wsnew.delete_rows(r,1)
                
def remove(sheet, row):
    for cell in row:
        if cell.value is None:
            sheet.delete_rows(row[0].row,1) 
        else:
            return
sheet=wsnew
for row in sheet:
    remove(sheet,row) 
wb.save("new.xlsx")
exit()






